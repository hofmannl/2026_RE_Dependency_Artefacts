from pathlib import Path
from openpyxl import load_workbook
from tqdm import tqdm

import argparse
import json
import time

START_ROW: int = 2
COMPARE_COL_1: str = "A"
COMPARE_COL_2: str = "B"

TRIM_WHITESPACE: bool = True
CASE_INSENSITIVE: bool = False
SKIP_IF_BOTH_EMPTY: bool = True


def normalize(v) -> str:
    if v is None:
        return ""
    s = str(v)
    if TRIM_WHITESPACE:
        s = s.strip()
    if CASE_INSENSITIVE:
        s = s.lower()
    return s

def cell_addr(col: str, row: int) -> str:
    return f"{col}{row}"

def load_word_pairs(path: str, sheet_name: str) -> list[tuple[str, str]]:
    """Load word pairs from a CSV file"""
    if isinstance(path, str):
        path = Path(path)
    
    results: list[tuple[str, str]] = []
    
    wb = load_workbook(path)
    for ws in tqdm(wb.worksheets, desc="Sheets", unit="sheet"):
        if ws.title.lower().strip() == sheet_name.lower().strip():
            for row in range(START_ROW, ws.max_row + 1):
                a = normalize(ws[cell_addr(COMPARE_COL_1, row)].value)
                b = normalize(ws[cell_addr(COMPARE_COL_2, row)].value)
                if SKIP_IF_BOTH_EMPTY and a == "" and b == "":
                    continue
                results.append((a, b))
            break  # Found the sheet, no need to continue
                
    return results 
        
    