from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook
from tqdm import tqdm

import argparse
import json
import time



INPUT_XLSX = Path(r"<USER_PATH>\automatic_sanity_checks\<dataset>.xlsx")
OUTPUT_XLSX = Path(r"<USER_PATH>\automatic_sanity_checks\<dataset>.xlsx")

SHEET_RULES: dict[str, int] = {
    "HS4ConOfE": 0,
    "HS4GenOfP": 0,
    "HS4GenOfA": 0,
    "HS4GenOfE": 0,
    "HS4SemOfP": 1,
    "HS4SemOfA": 1,
    "HS4SemOfE": 1,
    "HS4ConOfP": 0,
    "HS4ConOfA": 0,
}

START_ROW = 2
COMPARE_COL_1 = "A"
COMPARE_COL_2 = "B"
TARGET_COLUMNS = ["D", "E", "F"]

USE_DEFAULT_FOR_OTHER_SHEETS = False
DEFAULT_EQUAL_VALUE = 1


TRIM_WHITESPACE = True
CASE_INSENSITIVE = False
SKIP_IF_BOTH_EMPTY = True



def normalize(v) -> str:
    if v is None:
        return ""
    s = str(v)
    if TRIM_WHITESPACE:
        s = s.strip()
    if CASE_INSENSITIVE:
        s = s.lower()
    return s


def cell_addr(col: str, row: int) -> str:
    return f"{col}{row}"


def start_message() -> None:
    print("=" * 50)
    print("Excel processing started")
    print(f"Input file : {INPUT_XLSX}")
    print(f"Output file: {OUTPUT_XLSX}")
    print(f"Start row  : {START_ROW}")
    print(f"Compare    : {COMPARE_COL_1} vs {COMPARE_COL_2}")
    print(f"Targets    : {', '.join(TARGET_COLUMNS)}")
    print("")


def end_message(number_changes: int) -> None:
    print("")
    print(f"Finished. Total changes: {number_changes}")
    print(f"Output written to: {OUTPUT_XLSX}")
    print("=" * 50)


def parse_sheet_rules(sheet_rules_str: str | None) -> dict[str, int]:
    """
    Accept JSON like: {"HS4ConOfE":0,"HS4SemOfP":1}
    """
    if not sheet_rules_str:
        return {}
    try:
        data = json.loads(sheet_rules_str)
    except json.JSONDecodeError as e:
        raise SystemExit(f"Invalid JSON in --sheet-rules: {e}") from e

    if not isinstance(data, dict):
        raise SystemExit("--sheet-rules must be a JSON object, e.g. '{\"Sheet1\":1}'")

    out: dict[str, int] = {}
    for k, v in data.items():
        if int(v) not in (0, 1):
            raise SystemExit(f"Sheet '{k}' has invalid value '{v}'. Only 0 or 1 allowed.")
        out[str(k)] = int(v)
    return out


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Process XLSX sheets: compare two columns and write 0/1 into target columns per sheet."
    )

    p.add_argument("-i", "--input", type=Path, help="Input .xlsx file")
    p.add_argument("-o", "--output", type=Path, help="Output .xlsx file")

    p.add_argument("--start-row", type=int, help="First row to process (default: 2)")
    p.add_argument("--compare-cols", nargs=2, metavar=("COL1", "COL2"), help="Two columns to compare, e.g. A B")
    p.add_argument("--targets", nargs="+", help="Target columns to write, e.g. D E F")

    p.add_argument(
        "--sheet-rules",
        type=str,
        help='JSON mapping sheet name -> equal_value (0/1). Example: \'{"HS4ConOfE":0,"HS4SemOfP":1}\'',
    )

    p.add_argument("--use-default-for-other-sheets", action="store_true",
                   help="Process sheets not listed in --sheet-rules using --default-equal")
    p.add_argument("--default-equal", type=int, choices=[0, 1], help="Equal value for other sheets (0/1)")

    p.add_argument("--sheet-delay", type=float, help="Delay in seconds after each processed sheet")
    p.add_argument("--row-delay", type=float, help="Delay in seconds after each processed row")

    p.add_argument("--trim", action="store_true", help="Trim whitespace before comparing")
    p.add_argument("--case-insensitive", action="store_true", help="Compare case-insensitively")
    p.add_argument("--no-skip-empty", action="store_true", help="Do NOT skip rows where both compare cells are empty")

    return p


def apply_cli_to_globals(args: argparse.Namespace) -> None:
    """
    Overwrite global variables from CLI args (only when provided).
    """
    global INPUT_XLSX, OUTPUT_XLSX
    global START_ROW, COMPARE_COL_1, COMPARE_COL_2, TARGET_COLUMNS
    global SHEET_RULES, USE_DEFAULT_FOR_OTHER_SHEETS, DEFAULT_EQUAL_VALUE
    global TRIM_WHITESPACE, CASE_INSENSITIVE, SKIP_IF_BOTH_EMPTY

    if args.input is not None:
        INPUT_XLSX = args.input
    if args.output is not None:
        OUTPUT_XLSX = args.output

    if args.start_row is not None:
        START_ROW = args.start_row

    if args.compare_cols is not None:
        COMPARE_COL_1, COMPARE_COL_2 = args.compare_cols

    if args.targets is not None:
        TARGET_COLUMNS = args.targets

    if args.sheet_rules is not None:
        SHEET_RULES = parse_sheet_rules(args.sheet_rules)

    if args.use_default_for_other_sheets:
        USE_DEFAULT_FOR_OTHER_SHEETS = True

    if args.default_equal is not None:
        DEFAULT_EQUAL_VALUE = args.default_equal

    if args.sheet_delay is not None:
        SHEET_DELAY_SECONDS = max(0.0, float(args.sheet_delay))
    if args.row_delay is not None:
        ROW_DELAY_SECONDS = max(0.0, float(args.row_delay))


    if args.trim:
        TRIM_WHITESPACE = True
    if args.case_insensitive:
        CASE_INSENSITIVE = True


    SKIP_IF_BOTH_EMPTY = not args.no_skip_empty



def main() -> None:
    start_message()

    wb = load_workbook(INPUT_XLSX)
    number_changes = 0

    for ws in tqdm(wb.worksheets, desc="Sheets", unit="sheet"):
        sheet_name = ws.title

        if sheet_name in SHEET_RULES:
            equal_value = int(SHEET_RULES[sheet_name])
        elif USE_DEFAULT_FOR_OTHER_SHEETS:
            equal_value = int(DEFAULT_EQUAL_VALUE)
        else:
            continue

        if equal_value not in (0, 1):
            raise ValueError(f"Sheet '{sheet_name}': equal_value must be 0 or 1, got {equal_value}")
        not_equal_value = 1 - equal_value

        time.sleep(0.2)

        for row in range(START_ROW, ws.max_row + 1):
            a = normalize(ws[cell_addr(COMPARE_COL_1, row)].value)
            b = normalize(ws[cell_addr(COMPARE_COL_2, row)].value)

            if SKIP_IF_BOTH_EMPTY and a == "" and b == "":
                continue

            if a != b :
                continue


            for col in TARGET_COLUMNS:
                addr = cell_addr(col, row)
                if ws[addr].value != equal_value:
                    ws[addr].value = equal_value
                    number_changes += 1



    wb.save(OUTPUT_XLSX)
    end_message(number_changes)




if __name__ == "__main__":
    parser = build_arg_parser()
    args = parser.parse_args()
    apply_cli_to_globals(args)
    main()