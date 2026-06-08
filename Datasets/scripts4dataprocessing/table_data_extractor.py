import os
import csv
import re

from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# Konfiguration (anpassen)


INPUT_XLSX = r"" # Inputdatei
OUTPUT_DIR = r"" # Outputordner
OUTPUT_BASE = ""                            # Basisname der CSVs (Sheetname wird angehängt) e.g. g03_labeled
#OUTPUT_BASE = "g03_wordnet"
DELIMITER = "|"
START_ROW = 1                                     # erste zu untersuchende Zeile (1-basiert)

# Welche 3 Spalten pro Sheet (Excel-Spaltenbuchstaben)
COLUMNS_PER_SHEET = {
    "HS4ConOfE": ("A", "B", "G"),
    "HS4GenOfP": ("A", "B", "F"),
    "HS4GenOfA": ("A", "B", "G"),
    "HS4GenOfE": ("A", "B", "G"),
    "HS4SemOfP": ("A", "B", "F"),
    "HS4SemOfA": ("A", "B", "G"),
    "HS4SemOfE": ("A", "B", "G")
}

# Wordnet
#COLUMNS_PER_SHEET = {
#    "HS4ConOfE": ("A", "B", "C"),
#    "HS4GenOfA": ("A", "B", "C"),
#    "HS4GenOfE": ("A", "B", "C"),
#    "HS4SemOfA": ("A", "B", "C"),
#    "HS4SemOfE": ("A", "B", "C")
#}


INCLUDE_HEADER = False

# Hilfsfunktionen
def sanitize_filename(name: str) -> str:
    name = re.sub(r"[\\/:*?\"<>|]+", "_", name).strip()
    return name or "sheet"

def ensure_output_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)

def extract_3_columns(ws, cols, start_row: int):
    # Spaltenbuchstaben -> Index (1-basiert)
    col_indices = [column_index_from_string(c) for c in cols]
    min_col = min(col_indices)
    max_col = max(col_indices)

    positions = [i - min_col for i in col_indices]

    data = []

    for row in tqdm(
        ws.iter_rows(min_row=start_row, min_col=min_col, max_col=max_col, values_only=True),
        total=max(0, (ws.max_row or 0) - start_row + 1),
    ):
        v1 = row[positions[0]]
        v2 = row[positions[1]]
        v3 = row[positions[2]]

        if v1 is None and v2 is None and v3 is None:
            continue

        data.append((v1, v2, v3))

    return data

# Hauptlogik
def main():
    # Delimiter behandeln
    delimiter = "\t" if DELIMITER == r"\t" else DELIMITER
    if len(delimiter) != 1:
        raise ValueError("DELIMITER muss genau 1 Zeichen sein (oder '\\t').")

    # Outputordner sicherstellen
    ensure_output_dir(OUTPUT_DIR)

    # Workbook laden
    wb = load_workbook(INPUT_XLSX, data_only=True, read_only=True)

    # Validierung
    for sheet in COLUMNS_PER_SHEET.keys():
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' existiert nicht. Vorhanden: {wb.sheetnames}")
        if sheet not in COLUMNS_PER_SHEET:
            raise ValueError(f"Für Sheet '{sheet}' fehlen Spalten in COLUMNS_PER_SHEET.")

    # Verarbeitung pro Sheet
    for sheet in COLUMNS_PER_SHEET.keys():
        ws = wb[sheet]
        cols = COLUMNS_PER_SHEET[sheet]

        rows = extract_3_columns(ws, cols, START_ROW)

        out_name = f"{OUTPUT_BASE}_{sanitize_filename(sheet)}.csv"
        out_path = os.path.join(OUTPUT_DIR, out_name)

        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=delimiter)
            if INCLUDE_HEADER:
                writer.writerow(cols)  # Header = Spaltenbuchstaben (A, D, G)
            writer.writerows(rows)

        print(f"[OK] {sheet}: {len(rows)} Zeilen -> {out_path}")


if __name__ == "__main__":
    main()
